# GisDjango/utils.py
import os
from openpyxl import Workbook, load_workbook
from django.utils import timezone

def log_stock_action_to_excel(user_name, action, plant_name, branch_name, old_qty, new_qty):
    file_path = 'stock_management_logs.xlsx'

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Logs"
        ws.append(["Người thực hiện", "Ngày giờ", "Hành động", "Sản phẩm", "Chi nhánh", "Số lượng cũ", "Số lượng mới"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        user_name,
        timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
        action,
        plant_name,
        branch_name,
        old_qty,
        new_qty
    ])

    # 3. Lưu file
    wb.save(file_path)