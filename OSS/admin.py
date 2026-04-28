from django.contrib import messages
from django.contrib.auth.admin import UserAdmin
from django.contrib.gis import admin
from django.db import transaction
from django.db.models import ProtectedError
from django.contrib.auth import update_session_auth_hash

from .models import Plant, Category, Order, StoreBranch, StoreStock, User, PlantImage
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.gis.geos import Point

from OSS.helper.forms_helper import get_model_fields_data, save_model_from_request
from .models.forms import StoreStockForm

import os
from openpyxl import Workbook, load_workbook
from django.utils import timezone


# ================= USER ADMIN =================
class CustomUserAdmin(UserAdmin):
    fieldsets = UserAdmin.fieldsets + (
        (None, {'fields': ('phone_number', 'address')}),
    )
    add_fieldsets = UserAdmin.add_fieldsets + (
        (None, {'fields': ('phone_number', 'address')}),
    )

admin.site.register(User, CustomUserAdmin)


# ================= INLINE =================
class PlantImageInline(admin.TabularInline):
    model = PlantImage
    extra = 3


@admin.register(Plant)
class PlantAdmin(admin.GISModelAdmin):
    list_display = ('name', 'price', 'get_total_stock', 'category', 'is_available')
    search_fields = ('name',)
    inlines = [PlantImageInline]

    def get_total_stock(self, obj):
        return obj.total_stock
    get_total_stock.short_description = "Tồn kho"


class StockInline(admin.TabularInline):
    model = StoreStock
    extra = 1
    autocomplete_fields = ['plant']


@admin.register(StoreBranch)
class BranchAdmin(admin.GISModelAdmin):
    list_display = ('name', 'address', 'phone', 'delivery_radius', 'is_active')
    list_filter = ('is_active',)
    inlines = [StockInline]


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'created_at')


# ================= DASHBOARD =================
@staff_member_required
def dashboard(request):
    return render(request, 'admin/dashboard.html', {
        'total_plants': Plant.objects.count(),
        'total_categories': Category.objects.count(),
        'total_orders': Order.objects.count(),
        'total_branches': StoreBranch.objects.count(),
        'recent_plants': Plant.objects.order_by('-id')[:5],
        'recent_orders': Order.objects.order_by('-created_at')[:5],
    })


# ================= PLANT =================
@staff_member_required
def plant_management(request):
    plants = Plant.objects.all().order_by('-id')
    return render(request, 'admin/plant_list.html', {'plants': plants})


@staff_member_required
def plant_edit(request, pk=None):
    fields_data, instance = get_model_fields_data(Plant, pk)

    if request.method == 'POST':
        try:
            plant = save_model_from_request(request, Plant, instance=instance)

            images = request.FILES.getlist('images')
            for img in images:
                PlantImage.objects.create(plant=plant, image=img)

            return redirect('admin_plant_list')

        except Exception as e:
            print("ERROR:", e)
    return render(request, 'admin/dynamic_form.html', {
        'fields': fields_data,
        'title': "Sửa Cây" if pk else "Thêm Cây"
    })


@staff_member_required
def delete_plant(request, pk):
    plant = get_object_or_404(Plant, pk=pk)
    try:
        plant.delete()
    except ProtectedError:
        plant.is_available = False
        plant.save()
    return redirect('admin_plant_list')


# ================= ORDER =================
@staff_member_required
def order_management(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'admin/order_list.html', {'orders': orders})


@staff_member_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        old_status = order.status

        if new_status:
            try:
                with transaction.atomic():

                    if new_status == 'CANCELLED' and old_status != 'CANCELLED':
                        items = order.items.all()

                        for item in items:
                            stock = StoreStock.objects.filter(plant=item.plant).first()
                            if stock:
                                stock.quantity += item.quantity
                                stock.save()
                    order.status = new_status
                    order.save()

            except Exception as e:
                messages.error(request, str(e))

    return render(request, 'admin/order_detail.html', {'order': order})


# ================= BRANCH =================
@staff_member_required
def branch_management(request):
    branches = StoreBranch.objects.all().order_by('-id')
    return render(request, 'admin/branch_list.html', {'branches': branches})


@staff_member_required
def branch_edit(request, pk=None):
    fields_data, instance = get_model_fields_data(StoreBranch, pk)

    if request.method == 'POST':
        try:
            branch = save_model_from_request(request, StoreBranch, instance=instance)

            # fix PointField (nếu frontend gửi lat,lng)
            location = request.POST.get("location")
            if location:
                lat, lng = map(float, location.split(","))
                branch.location = Point(lng, lat)
                branch.save()

            return redirect('admin_branch_list')

        except Exception as e:
            print(e)

    return render(request, 'admin/dynamic_form.html', {
        'fields': fields_data,
        'title': "Sửa Chi Nhánh" if pk else "Thêm Chi Nhánh"
    })


@staff_member_required
def branch_delete(request, pk):
    branch = get_object_or_404(StoreBranch, pk=pk)
    try:
        branch.delete()
    except ProtectedError:
        messages.error(request, 'Không thể xóa')
    return redirect('admin_branch_list')


# ================= CATEGORY =================
@staff_member_required
def category_management(request):
    categories = Category.objects.all().order_by('-id')
    return render(request, 'admin/category_list.html', {'categories': categories})


@staff_member_required
def category_edit(request, pk=None):
    fields_data, instance = get_model_fields_data(Category, pk)

    if request.method == 'POST':
        save_model_from_request(request, Category, instance=instance)
        return redirect('admin_category_list')

    return render(request, 'admin/dynamic_form.html', {
        'fields': fields_data,
        'title': "Sửa Danh Mục" if pk else "Thêm Danh Mục"
    })

def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    try:
        category.delete()
    except ProtectedError:
        messages.error(request, 'Không thể xóa')
    return redirect('admin_category_list')


# ================= STOCK =================
@staff_member_required
def stock_management(request):
    stocks = StoreStock.objects.select_related('branch', 'plant')
    return render(request, 'admin/stock_list.html', {'stocks': stocks})


@staff_member_required
def stock_edit(request, pk=None):
    fields_data, instance = get_model_fields_data(StoreStock, pk)

    # Lưu lại số lượng cũ trước khi POST
    old_qty = instance.quantity if instance else 0
    action_type = "CẬP NHẬT" if pk else "NHẬP MỚI"

    if request.method == 'POST':
        # Lưu model bằng helper của ông
        stock = save_model_from_request(request, StoreStock, instance=instance)

        # Ghi log Excel sau khi lưu thành công
        log_stock_to_excel(
            user_name=request.user.username,
            action=action_type,
            plant_name=stock.plant.name,
            branch_name=stock.branch.name,
            old_qty=old_qty,
            new_qty=stock.quantity
        )

        messages.success(request, f"{action_type} kho thành công!")
        return redirect('admin_stock_list')

    return render(request, 'admin/dynamic_form.html', {
        'fields': fields_data,
        'title': "Cập nhật kho" if pk else "Nhập hàng mới"
    })


@staff_member_required
def stock_delete(request, pk):
    stock = get_object_or_404(StoreStock, pk=pk)

    # Lưu thông tin trước khi xóa
    u_name = request.user.username
    p_name = stock.plant.name
    b_name = stock.branch.name
    q_old = stock.quantity

    stock.delete()

    # Ghi log hành động xóa
    log_stock_to_excel(u_name, "XÓA DỮ LIỆU KHO", p_name, b_name, q_old, 0)

    messages.warning(request, "Đã xóa dữ liệu kho và ghi nhật ký.")
    return redirect('admin_stock_list')


# ================= USER =================
@staff_member_required
def user_management(request):
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'admin/user_list.html', {'users': users})


@staff_member_required
def user_edit(request, pk=None):
    user_obj = get_object_or_404(User, pk=pk) if pk else None
    fields_data, _ = get_model_fields_data(User, pk)

    if request.method == 'POST':
        user = user_obj or User()

        user.username = request.POST.get('username')
        user.email = request.POST.get('email')

        password = request.POST.get('password')
        if password:
            user.set_password(password)

        user.save()

        if password and request.user.id == user.id:
            update_session_auth_hash(request, user)

        return redirect('admin_user_list')

    return render(request, 'admin/dynamic_form.html', {
        'fields': fields_data,
        'title': "Sửa User" if pk else "Thêm User"
    })


@staff_member_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)

    if request.user.id == user.id:
        messages.error(request, "Không thể tự khóa")
        return redirect('admin_user_list')

    user.is_active = not user.is_active
    user.save()

    return redirect('admin_user_list')


# ================= ADMIN STOCK =================
@admin.register(StoreStock)
class StoreStockAdmin(admin.ModelAdmin):
    form = StoreStockForm
    list_display = ('branch', 'plant', 'quantity')
    list_filter = ('branch', 'plant')


def log_stock_to_excel(user_name, action, plant_name, branch_name, old_qty, new_qty):
    file_path = 'stock_management_logs.xlsx'

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Logs"
        ws.append(["Người thực hiện", "Ngày giờ", "Hành động", "Sản phẩm", "Chi nhánh", "Số lượng cũ", "Số lượng mới"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        user_name,
        timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
        action,
        plant_name,
        branch_name,
        old_qty,
        new_qty
    ])
    wb.save(file_path)